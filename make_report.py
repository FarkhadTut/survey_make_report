from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import os
import pandas as pd 
from openpyxl.styles import Alignment, Font
from glob import glob
from utils.utils import strip_tags

SHEET_NAME_DB = 'db'
DROP_COLS = ['deviceid',
             'audit',
             'audit_URL',
             'cur_time',
             'cur_date',
             'cur_datetime',
             'cur_time', 
             'cur_date', 
             'cur_datetime', 
             'mahalla_code', 
             'district', 
             '_id', 
             '_uuid', 
             '_submission_time', 
             '_status', 
             '__version__', 
             '_index', 
             '_validation_status',
             '_notes',
             '_submitted_by',
             '_tags', 
            '1.2. Интервьюерни танланг:',
            'respondents',
             ]




new_sheet = 'Свод'

SPACE_BETWEEN = 5




def save(wb, filename_out):
    wb.save(filename_out)
    wb.close()
# pivot_table = df.pivot_table(values='_id', index='1. Туманингизни танланг?', columns='14. Маълумотингиз?', aggfunc='count')

def move_other_to_end(lst):
    popped = []
    for item in lst:
        if 'бошқа' in item.lower():
            lst.pop(lst.index(item))
            popped.append(item)

    lst = lst + popped
    return lst

def create_new_sheet(filename, new_sheet):
    start_col = '5. Корхонанинг асосий иқтисодий фаолият тури:'
    end_col = '63. Тадбиркорлик муҳити ва бизнес шароитларини яхшилаш учун таклифларингиз:'
    DROP_COLS.append(start_col)
    wb = openpyxl.load_workbook(filename)
    df = pd.read_excel(filename)
    regions = df[start_col].unique().tolist()
    wb.create_sheet(new_sheet)
    ws = wb[new_sheet]
    
    ws_db = wb.active
    ws_db.title = SHEET_NAME_DB
    max_rows = ws_db.max_row
    max_columns = ws_db.max_column

    cur_col = 1
    columns = []
    
    survey_file = glob(os.path.join('survey', '*.xlsx'))[0]
    df_survey = pd.read_excel(survey_file)

    label_col = [c for c in df_survey.columns.values if c.startswith('label')][0]

    MULTIPLE_QUESTIONS = df_survey[df_survey['type'].fillna('nan').str.startswith('select_multiple ')][label_col]
    MULTIPLE_QUESTIONS = [strip_tags(c) for c in MULTIPLE_QUESTIONS]
    
    note_cols = df_survey[df_survey['type'].str.strip() == 'note'][label_col]
    note_cols = [strip_tags(c) for c in note_cols]

    text_cols = df_survey[df_survey['type'].str.strip() == 'text'][label_col]
    text_cols = [strip_tags(c) for c in text_cols]

    DROP_COLS.extend(note_cols)
    DROP_COLS.extend(text_cols)

    passed_start_col = False
    for col_i in range(1, max_columns+1):
        column = ws_db[f'{get_column_letter(col_i)}1'].value
        if column.strip() == start_col:
            passed_start_col = True
        if not passed_start_col:
            continue
        if column not in DROP_COLS \
            and '(Бошқа' not in column \
            and not column.strip().endswith(')***')  \
            and column not in MULTIPLE_QUESTIONS:
            
            answers = sorted([str(c) for c in df[column].dropna().unique()])
            answers = move_other_to_end(answers)
            answers_count = len(answers)
            cur_col += 1
            ws[f'{get_column_letter(cur_col)}1'] = column
            ws[f'{get_column_letter(cur_col)}1'].alignment = Alignment(wrap_text=True, 
                                                                       horizontal='center',
                                                                       vertical='center')
            
            if not(df[column].dtype in ['int64', 'float64'] or answers_count > 25):
                #Single choice question
                start_col = cur_col
                
                for i, answer in enumerate(answers):
                    ws[f'{get_column_letter(cur_col)}2'] = answer
                    ws[f'{get_column_letter(cur_col)}2'].alignment = Alignment(wrap_text=True, 
                                                                                horizontal='center',
                                                                                vertical='center')
                    if i != answers_count-1:
                        cur_col += 1 
                    
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=cur_col)
           

           
           ################ Collecting columns #######################################################
                columns.append({'question': column,
                                'col': start_col,
                                'end_col': cur_col,
                                'col_db': col_i,
                                'answers': answers,
                                'type': 'single'})

            elif df[column].dtype in ['int64', 'float64']:
                if df[column].max == 1:
                    columns.append({'question':column,
                                    'col_db': col_i,
                                    'col': cur_col,
                                    'type': 'multiple'})
                else:
                    columns.append({'question':column,
                                    'col_db': col_i,
                                    'col': cur_col,
                                    'type': 'numeric'})
                    

                ws.merge_cells(start_row=1, start_column=cur_col, end_row=2, end_column=cur_col)
                
                
        ##############################################################################
        elif column == start_col:
            cur_col += 1
            ws[f'{get_column_letter(cur_col)}1'] = column
            ws[f'{get_column_letter(cur_col)}1'].alignment = Alignment(vertical='center',
                                                                       horizontal='center',
                                                                       wrap_text=True)
            regions_col_db = col_i
            columns.append({'question':column,
                            'col_db': col_i,
                            'col': cur_col,
                            'type': 'regions'})

    for i, region in enumerate(regions):
        ws[f'A{i+3}'] = region
        ws[f'A{i+3+len(regions) + SPACE_BETWEEN}'] = region

    ws.row_dimensions[1].height = 172.5
    ws.row_dimensions[2].height = 112.5
    ws.column_dimensions['A'].width = 30
    return wb, columns, regions_col_db, regions


def calculate(wb, columns, regions_col_db, regions):

    AVERAGE =[
        '7. Бугунги кунда корхона қандай қувватда ишламоқда? (мавжуд қувватлардан фойдаланиш даражаси)',
        '9. Фикрингизча юқоридаги муаммолар ечилса, ишлаб чиқариш ҳажмини неча фоизга оширса бўлади?',
         ]

    regions_col_db = get_column_letter(regions_col_db)
    ws_db = wb.active
    sheet_name_db = ws_db.title 
    ws = wb[new_sheet]
    start_row_idx = 3
    end_row_idx = start_row_idx + len(regions)
    for row in range(start_row_idx, end_row_idx):
        for column in columns:
            question = column['question']
            region = f'$A{row}'
            col_svod = get_column_letter(column['col'])
            col_db = get_column_letter(column['col_db'])
            if column['type'] == 'numeric' or column['type'] == 'multiple':
                formula = f"""=SUMIF('{sheet_name_db}'!${regions_col_db}:${regions_col_db}, '{new_sheet}'!{region}, '{sheet_name_db}'!{col_db}:{col_db})"""
                if question in AVERAGE:
                    formula = formula.replace('SUMIF', 'AVERAGEIF')
                ws[f'{col_svod}{row}'] = formula
            elif column['type'] == 'single':
                for col_svod_single in range(column['col'], column['end_col']+1):
                    col_svod_single = get_column_letter(col_svod_single)
                    formula = f"""=COUNTIFS('{sheet_name_db}'!${regions_col_db}:${regions_col_db}, '{new_sheet}'!{region}, '{sheet_name_db}'!${col_db}:${col_db}, '{new_sheet}'!{col_svod_single}${2})"""
                    ws[f'{col_svod_single}{row}'] = formula
            
            elif column['type'] == 'regions':
                formula = f"""=COUNTIF('{sheet_name_db}'!${regions_col_db}:${regions_col_db}, '{new_sheet}'!{region})"""
                ws[f'{col_svod}{row}'] = formula

        
    ###### add totals #####
    for column in columns:
        if column['type'] == 'numeric' or column['type'] == 'multiple' or column['type'] == 'regions':
            col_svod = get_column_letter(column['col'])
            ws[f'{col_svod}{end_row_idx}'] = f'=SUM({col_svod}{start_row_idx}:{col_svod}{end_row_idx-1})'
            ws[f'{col_svod}{end_row_idx}'].font = Font(bold=True)

        else:
            for col_svod_single in range(column['col'], column['end_col']+1):
                col_svod_single = get_column_letter(col_svod_single)
                ws[f'{col_svod_single}{end_row_idx}'] = f'=SUM({col_svod_single}{start_row_idx}:{col_svod_single}{end_row_idx-1})'
                ws[f'{col_svod_single}{end_row_idx}'].font = Font(bold=True)


        ### works but no need actually, calculates average of averages but in the right way
        # if column['question'] in AVERAGE:
        #     col_avg_db = get_column_letter(column['col_db']) 
        #     col_avg = get_column_letter(column['col'])
        #     ws[f'{col_avg}17'] = f'=AVERAGE({SHEET_NAME_DB}!{col_avg_db}:{col_avg_db})'
        #     ws[f'{col_avg}17'].font = Font(bold=True)
        ####################################################################################

    #####################

    
    one_to_one_step = SPACE_BETWEEN + len(regions)
    start_row_idx = end_row_idx + SPACE_BETWEEN
    end_row_idx = start_row_idx + len(regions)

    ##### percentage count #######
    for row in range(start_row_idx, end_row_idx+1):
        for column in columns:
            if column['type'] == 'numeric' or column['type'] == 'multiple' or column['type'] == 'regions':
                col_svod = get_column_letter(column['col'])
                ws[f'{col_svod}{row}'] = f'={col_svod}{row-one_to_one_step}/$B{row-one_to_one_step}'
                ws[f'{col_svod}{end_row_idx}'].font = Font(bold=True)
                ws[f'{col_svod}{row}'].number_format = '0.0%'

            else:
                for col_svod_single in range(column['col'], column['end_col']+1):
                    col_svod_single = get_column_letter(col_svod_single)
                    ws[f'{col_svod_single}{row}'] = f'={col_svod_single}{row-one_to_one_step}/$B{row-one_to_one_step}'
                    ws[f'{col_svod_single}{end_row_idx}'].font = Font(bold=True)
                    ws[f'{col_svod_single}{row}'].number_format = '0.0%'
    ####################################


            

            

        

    return wb




def make_report(db_filename):
    db_filename = os.path.join('data', db_filename)
    filename_out = db_filename#f"out\\freq\\output_{os.path.basename(db_filename)}"
    wb, columns, regions_col_db, regions = create_new_sheet(db_filename, new_sheet=new_sheet)
    wb = calculate(wb, columns, regions_col_db, regions)
    save(wb, filename_out=filename_out)
    print("Massive Frequency table ready!....")







make_report('db_2024_07_29.xlsx')